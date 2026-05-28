# -*- coding: utf-8 -*-
from __future__ import annotations

import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import io
import math
import os
import re
import shutil
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import cv2
import numpy as np
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent
CONFIG_TXT = BASE_DIR / "采集配置.txt"
INPUT_XLSX = BASE_DIR / "待采集视频链接.xlsx"
OUTPUT_XLSX = BASE_DIR / "采集结果.xlsx"
SCREENSHOT_DIR = BASE_DIR / "screenshots"
MODEL_DIR = BASE_DIR / "models"
YUNET_MODEL = MODEL_DIR / "face_detection_yunet_2023mar.onnx"

MAX_CONCURRENT = 3
RETRY_CONCURRENT = 1
MAX_RETRY_ROUNDS = 4
RESOLVE_CONCURRENT = 8
VIEWPORT = {"width": 1280, "height": 720}
CHINA_TZ = timezone(timedelta(hours=8))

CONFIG_TEMPLATE = """# 抖音采集脚本配置
# 格式：键=数字。改完保存后重新运行 python collect_douyin.py

# 一次同时打开多少个视频页面。越大越快，但也越容易触发验证码。
一次性采集数量=3

# 失败/验证码重试时的并发数，建议保持 1。
失败重试一次性采集数量=1

# 浏览器窗口大小。截图最大不会超过这个窗口。
浏览器宽度=1280
浏览器高度=720

# 可选：手动指定浏览器路径。不填则自动查找常见浏览器，找不到则尝试使用随包 Chromium。
# Windows 示例：C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe
# macOS 示例：/Applications/Google Chrome.app
浏览器路径=

# 截图裁剪参数，单位像素。
# 下面几个值以“主视频区域”为基准向外扩展：
# 左扩展越大，截图越往左；右扩展越大，越容易包含右侧推荐栏。
截图左扩展=160
截图右扩展=320
截图上扩展=80
截图下扩展=220

# 截图最小宽度。想更窄可以调小，例如 900；想更宽可以调大，例如 1200。
截图最小宽度=1120

# 截图最大宽度/高度。0 表示不限制，只受浏览器窗口限制。
截图最大宽度=0
截图最大高度=0

# Excel 里嵌入图片的显示尺寸，不影响 screenshots 文件夹里的原图。
Excel图片最大宽度=320
Excel图片最大高度=205
"""

INPUT_HEADER = "抖音链接，支持带口令的文本"
OUTPUT_HEADERS = [
    "对接人",
    "账号昵称",
    "账号ID",
    "粉丝量(w)",
    "视频链接",
    "点赞量",
    "评论量",
    "转发",
    "发布时间",
    "发布截图",
    "备注",
]

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)

YUNET_URLS = [
    "https://github.com/opencv/opencv_zoo/raw/main/models/face_detection_yunet/face_detection_yunet_2023mar.onnx",
    "https://raw.githubusercontent.com/opencv/opencv_zoo/main/models/face_detection_yunet/face_detection_yunet_2023mar.onnx",
]


def clamp_int(value: int, minimum: int, maximum: int) -> int:
    return max(minimum, min(maximum, value))


def read_config_entries() -> dict[str, str]:
    if not CONFIG_TXT.exists():
        CONFIG_TXT.write_text(CONFIG_TEMPLATE, encoding="utf-8-sig")

    text = CONFIG_TXT.read_text(encoding="utf-8-sig")
    if "浏览器路径" not in text:
        with CONFIG_TXT.open("a", encoding="utf-8-sig") as file:
            file.write(
                "\n# 可选：手动指定浏览器 exe 路径。不填则自动查找常见浏览器。\n"
                "浏览器路径=\n"
            )
        text = CONFIG_TXT.read_text(encoding="utf-8-sig")

    entries: dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, raw_value = line.split("=", 1)
        key = key.strip()
        value = raw_value.strip().strip('"').strip("'")
        if key:
            entries[key] = value
    return entries


def parse_config_file() -> dict[str, int]:
    entries = read_config_entries()
    values: dict[str, int] = {}
    for key, raw_value in entries.items():
        value_match = re.search(r"-?\d+", raw_value)
        if not value_match:
            continue
        values[key] = int(value_match.group(0))
    return values


def load_settings() -> Settings:
    entries = read_config_entries()
    values = parse_config_file()
    return Settings(
        max_concurrent=clamp_int(values.get("一次性采集数量", MAX_CONCURRENT), 1, 10),
        retry_concurrent=clamp_int(values.get("失败重试一次性采集数量", RETRY_CONCURRENT), 1, 10),
        viewport_width=clamp_int(values.get("浏览器宽度", VIEWPORT["width"]), 800, 3840),
        viewport_height=clamp_int(values.get("浏览器高度", VIEWPORT["height"]), 500, 2160),
        crop_left_extra=clamp_int(values.get("截图左扩展", 160), 0, 4000),
        crop_right_extra=clamp_int(values.get("截图右扩展", 320), 0, 4000),
        crop_top_extra=clamp_int(values.get("截图上扩展", 80), 0, 4000),
        crop_bottom_extra=clamp_int(values.get("截图下扩展", 220), 0, 4000),
        crop_min_width=clamp_int(values.get("截图最小宽度", 1120), 1, 10000),
        crop_max_width=clamp_int(values.get("截图最大宽度", 0), 0, 10000),
        crop_max_height=clamp_int(values.get("截图最大高度", 0), 0, 10000),
        excel_image_max_width=clamp_int(values.get("Excel图片最大宽度", 320), 50, 2000),
        excel_image_max_height=clamp_int(values.get("Excel图片最大高度", 205), 50, 2000),
        browser_path=entries.get("浏览器路径", ""),
    )


@dataclass
class InputItem:
    row: int
    raw_text: str
    url: str
    page_url: str | None = None


@dataclass
class CollectResult:
    contact: str = ""
    nickname: str = ""
    account_id: str = ""
    fans_w: float | None = None
    video_link: str = ""
    digg_count: int | None = None
    comment_count: int | None = None
    share_count: int | None = None
    publish_time: str = ""
    screenshot_path: Path | None = None
    remark: str = ""


@dataclass(frozen=True)
class Settings:
    max_concurrent: int = 3
    retry_concurrent: int = 1
    viewport_width: int = 1280
    viewport_height: int = 720
    crop_left_extra: int = 160
    crop_right_extra: int = 320
    crop_top_extra: int = 80
    crop_bottom_extra: int = 220
    crop_min_width: int = 1120
    crop_max_width: int = 0
    crop_max_height: int = 0
    excel_image_max_width: int = 320
    excel_image_max_height: int = 205
    browser_path: str = ""


class FaceDetector:
    def __init__(self, yunet_model: Path | None):
        self.yunet = None
        if yunet_model and yunet_model.exists():
            try:
                self.yunet = cv2.FaceDetectorYN_create(
                    str(yunet_model),
                    "",
                    (320, 320),
                    score_threshold=0.6,
                    nms_threshold=0.3,
                    top_k=5000,
                )
            except Exception:
                self.yunet = None

        self.haar_frontal = self._load_cascade("haarcascade_frontalface_alt2.xml")
        self.haar_profile = self._load_cascade("haarcascade_profileface.xml")

    def _load_cascade(self, filename: str) -> cv2.CascadeClassifier:
        haar_dir = Path(cv2.data.haarcascades)
        source = haar_dir / filename
        if not source.exists():
            return cv2.CascadeClassifier()

        try:
            str(source).encode("ascii")
            cascade_path = source
        except UnicodeEncodeError:
            temp_dir = Path(tempfile.gettempdir()) / "douyin_collector_cv2"
            temp_dir.mkdir(parents=True, exist_ok=True)
            cascade_path = temp_dir / filename
            if not cascade_path.exists() or cascade_path.stat().st_size != source.stat().st_size:
                shutil.copyfile(source, cascade_path)

        cascade = cv2.CascadeClassifier(str(cascade_path))
        if cascade.empty():
            return cv2.CascadeClassifier()
        return cascade

    def has_face(self, bgr_image: np.ndarray) -> bool:
        if bgr_image is None or bgr_image.size == 0:
            return False

        if self.yunet is not None:
            height, width = bgr_image.shape[:2]
            self.yunet.setInputSize((width, height))
            _, faces = self.yunet.detect(bgr_image)
            if faces is not None and len(faces) > 0:
                return True

        gray = cv2.cvtColor(bgr_image, cv2.COLOR_BGR2GRAY)
        gray = cv2.equalizeHist(gray)
        for cascade in (self.haar_frontal, self.haar_profile):
            if cascade.empty():
                continue
            faces = cascade.detectMultiScale(
                gray,
                scaleFactor=1.08,
                minNeighbors=5,
                minSize=(36, 36),
            )
            if len(faces) > 0:
                return True
        return False


def ensure_input_workbook() -> bool:
    if INPUT_XLSX.exists():
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "待采集视频链接"
    ws["A1"] = INPUT_HEADER
    ws.column_dimensions["A"].width = 90
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(vertical="center")
    wb.save(INPUT_XLSX)
    return True


def extract_first_url(text: str) -> str | None:
    if not text:
        return None
    match = re.search(r"https?://[^\s，,。；;）)】\]]+", text)
    if not match:
        return None
    return match.group(0).rstrip(".,;，。；、")


def is_douyin_url(url: str) -> bool:
    return any(domain in url.lower() for domain in ("douyin.com", "iesdouyin.com"))


def read_input_items() -> list[InputItem]:
    ensure_input_workbook()
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active
    items: list[InputItem] = []
    for row in range(1, ws.max_row + 1):
        raw = ws.cell(row=row, column=1).value
        if raw is None:
            continue
        raw_text = str(raw).strip()
        url = extract_first_url(raw_text)
        if not url:
            continue
        if not is_douyin_url(url):
            continue
        items.append(InputItem(row=row, raw_text=raw_text, url=url))
    return items


def canonical_video_url_from_text(text: str) -> str | None:
    match = re.search(r"(?:/share)?/video/(\d+)", text)
    if match:
        return f"https://www.douyin.com/video/{match.group(1)}"
    return None


def resolve_douyin_video_url(url: str) -> str | None:
    direct = canonical_video_url_from_text(url)
    if direct:
        return direct

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9",
    }
    current = url
    for _ in range(6):
        direct = canonical_video_url_from_text(current)
        if direct:
            return direct
        try:
            response = requests.get(
                current,
                allow_redirects=False,
                headers=headers,
                timeout=12,
            )
        except Exception:
            return None

        location = response.headers.get("Location") or response.headers.get("location")
        if location:
            current = urljoin(current, location)
            continue

        direct = canonical_video_url_from_text(response.text[:3000])
        if direct:
            return direct
        break
    return canonical_video_url_from_text(current)


def prepare_input_items(items: list[InputItem]) -> list[InputItem]:
    if not items:
        return items

    print(f"先解析 {len(items)} 条短链，减少浏览器跳转失败。")
    resolved: list[InputItem] = list(items)

    def resolve_one(index: int, item: InputItem) -> tuple[int, str | None]:
        for attempt in range(3):
            page_url = resolve_douyin_video_url(item.url)
            if page_url:
                return index, page_url
            if attempt < 2:
                time_to_sleep = 0.5 * (attempt + 1)
                time.sleep(time_to_sleep)
        return index, None

    workers = min(RESOLVE_CONCURRENT, max(1, len(items)))
    ok = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(resolve_one, idx, item) for idx, item in enumerate(items)]
        for future in as_completed(futures):
            idx, page_url = future.result()
            item = items[idx]
            if page_url:
                ok += 1
                resolved[idx] = InputItem(
                    row=item.row,
                    raw_text=item.raw_text,
                    url=item.url,
                    page_url=page_url,
                )
            else:
                resolved[idx] = item
    print(f"短链解析完成：{ok}/{len(items)} 条已解析到直达视频页。")
    return resolved


def read_existing_contacts() -> dict[str, str]:
    if not OUTPUT_XLSX.exists():
        return {}
    try:
        wb = load_workbook(OUTPUT_XLSX)
    except Exception:
        return {}
    ws = wb.active
    contacts: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        link = ws.cell(row=row, column=5).value
        contact = ws.cell(row=row, column=1).value
        if link and contact:
            contacts[str(link).strip()] = str(contact).strip()
    return contacts


def ensure_yunet_model() -> Path | None:
    if YUNET_MODEL.exists() and YUNET_MODEL.stat().st_size > 100_000:
        return YUNET_MODEL

    MODEL_DIR.mkdir(exist_ok=True)
    for url in YUNET_URLS:
        try:
            response = requests.get(url, timeout=25)
            response.raise_for_status()
            if len(response.content) > 100_000:
                YUNET_MODEL.write_bytes(response.content)
                return YUNET_MODEL
        except Exception:
            continue
    return None


def browser_engine_for_path(path: Path) -> str:
    lowered_parts = {part.lower() for part in path.parts}
    if path.name.lower() in {"firefox", "firefox.exe"} or "firefox.app" in lowered_parts:
        return "firefox"
    return "chromium"


def app_bundle_executable(app_path: Path) -> Path | None:
    if sys.platform != "darwin" or app_path.suffix.lower() != ".app":
        return None

    macos_dir = app_path / "Contents" / "MacOS"
    if not macos_dir.exists():
        return None

    preferred = macos_dir / app_path.stem
    if preferred.exists():
        return preferred

    for child in macos_dir.iterdir():
        if child.is_file() and not child.name.startswith("."):
            return child
    return None


def browser_candidates() -> list[tuple[str, str, Path]]:
    if sys.platform == "darwin":
        applications = Path("/Applications")
        user_applications = Path.home() / "Applications"
        raw_candidates = [
            ("Google Chrome", "chromium", applications / "Google Chrome.app" / "Contents" / "MacOS" / "Google Chrome"),
            ("Google Chrome", "chromium", user_applications / "Google Chrome.app" / "Contents" / "MacOS" / "Google Chrome"),
            ("Microsoft Edge", "chromium", applications / "Microsoft Edge.app" / "Contents" / "MacOS" / "Microsoft Edge"),
            ("Microsoft Edge", "chromium", user_applications / "Microsoft Edge.app" / "Contents" / "MacOS" / "Microsoft Edge"),
            ("Brave", "chromium", applications / "Brave Browser.app" / "Contents" / "MacOS" / "Brave Browser"),
            ("Brave", "chromium", user_applications / "Brave Browser.app" / "Contents" / "MacOS" / "Brave Browser"),
            ("Arc", "chromium", applications / "Arc.app" / "Contents" / "MacOS" / "Arc"),
            ("Arc", "chromium", user_applications / "Arc.app" / "Contents" / "MacOS" / "Arc"),
            ("Opera", "chromium", applications / "Opera.app" / "Contents" / "MacOS" / "Opera"),
            ("Opera", "chromium", user_applications / "Opera.app" / "Contents" / "MacOS" / "Opera"),
            ("Opera GX", "chromium", applications / "Opera GX.app" / "Contents" / "MacOS" / "Opera GX"),
            ("Opera GX", "chromium", user_applications / "Opera GX.app" / "Contents" / "MacOS" / "Opera GX"),
            ("Vivaldi", "chromium", applications / "Vivaldi.app" / "Contents" / "MacOS" / "Vivaldi"),
            ("Vivaldi", "chromium", user_applications / "Vivaldi.app" / "Contents" / "MacOS" / "Vivaldi"),
            ("Chromium", "chromium", applications / "Chromium.app" / "Contents" / "MacOS" / "Chromium"),
            ("Chromium", "chromium", user_applications / "Chromium.app" / "Contents" / "MacOS" / "Chromium"),
            ("QQ Browser", "chromium", applications / "QQBrowser.app" / "Contents" / "MacOS" / "QQBrowser"),
            ("QQ Browser", "chromium", applications / "QQ浏览器.app" / "Contents" / "MacOS" / "QQ浏览器"),
            ("360 Browser", "chromium", applications / "360极速浏览器.app" / "Contents" / "MacOS" / "360极速浏览器"),
            ("Sogou Browser", "chromium", applications / "搜狗浏览器.app" / "Contents" / "MacOS" / "搜狗浏览器"),
            ("Firefox", "firefox", applications / "Firefox.app" / "Contents" / "MacOS" / "firefox"),
            ("Firefox", "firefox", user_applications / "Firefox.app" / "Contents" / "MacOS" / "firefox"),
        ]
    else:
        program_files = Path(os.environ.get("PROGRAMFILES", r"C:\Program Files"))
        program_files_x86 = Path(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"))
        local_app_data = Path(os.environ.get("LOCALAPPDATA", str(Path.home() / "AppData" / "Local")))

        raw_candidates = [
            ("Google Chrome", "chromium", program_files / "Google" / "Chrome" / "Application" / "chrome.exe"),
            ("Google Chrome", "chromium", program_files_x86 / "Google" / "Chrome" / "Application" / "chrome.exe"),
            ("Microsoft Edge", "chromium", program_files / "Microsoft" / "Edge" / "Application" / "msedge.exe"),
            ("Microsoft Edge", "chromium", program_files_x86 / "Microsoft" / "Edge" / "Application" / "msedge.exe"),
            ("Brave", "chromium", program_files / "BraveSoftware" / "Brave-Browser" / "Application" / "brave.exe"),
            ("Brave", "chromium", program_files_x86 / "BraveSoftware" / "Brave-Browser" / "Application" / "brave.exe"),
            ("Brave", "chromium", local_app_data / "BraveSoftware" / "Brave-Browser" / "Application" / "brave.exe"),
            ("Opera", "chromium", program_files / "Opera" / "launcher.exe"),
            ("Opera", "chromium", program_files_x86 / "Opera" / "launcher.exe"),
            ("Opera", "chromium", local_app_data / "Programs" / "Opera" / "launcher.exe"),
            ("Opera GX", "chromium", local_app_data / "Programs" / "Opera GX" / "launcher.exe"),
            ("Vivaldi", "chromium", program_files / "Vivaldi" / "Application" / "vivaldi.exe"),
            ("Vivaldi", "chromium", program_files_x86 / "Vivaldi" / "Application" / "vivaldi.exe"),
            ("Vivaldi", "chromium", local_app_data / "Vivaldi" / "Application" / "vivaldi.exe"),
            ("Chromium", "chromium", program_files / "Chromium" / "Application" / "chrome.exe"),
            ("Chromium", "chromium", program_files_x86 / "Chromium" / "Application" / "chrome.exe"),
            ("Chromium", "chromium", local_app_data / "Chromium" / "Application" / "chrome.exe"),
            ("360 Chrome", "chromium", program_files / "360" / "360Chrome" / "Chrome" / "Application" / "360chrome.exe"),
            ("360 Chrome", "chromium", program_files_x86 / "360" / "360Chrome" / "Chrome" / "Application" / "360chrome.exe"),
            ("360 Chrome", "chromium", local_app_data / "360Chrome" / "Chrome" / "Application" / "360chrome.exe"),
            ("360 Safe Browser", "chromium", program_files / "360" / "360se6" / "Application" / "360se.exe"),
            ("360 Safe Browser", "chromium", program_files_x86 / "360" / "360se6" / "Application" / "360se.exe"),
            ("QQ Browser", "chromium", program_files / "Tencent" / "QQBrowser" / "QQBrowser.exe"),
            ("QQ Browser", "chromium", program_files_x86 / "Tencent" / "QQBrowser" / "QQBrowser.exe"),
            ("QQ Browser", "chromium", local_app_data / "Tencent" / "QQBrowser" / "QQBrowser.exe"),
            ("Sogou Explorer", "chromium", program_files / "SogouExplorer" / "SogouExplorer.exe"),
            ("Sogou Explorer", "chromium", program_files_x86 / "SogouExplorer" / "SogouExplorer.exe"),
            ("Sogou Explorer", "chromium", local_app_data / "SogouExplorer" / "SogouExplorer.exe"),
            ("Liebao Browser", "chromium", program_files / "liebao" / "liebao.exe"),
            ("Liebao Browser", "chromium", program_files_x86 / "liebao" / "liebao.exe"),
            ("Maxthon", "chromium", program_files / "Maxthon" / "Application" / "Maxthon.exe"),
            ("Maxthon", "chromium", program_files_x86 / "Maxthon" / "Application" / "Maxthon.exe"),
            ("Maxthon", "chromium", local_app_data / "Maxthon" / "Application" / "Maxthon.exe"),
            ("Firefox", "firefox", program_files / "Mozilla Firefox" / "firefox.exe"),
            ("Firefox", "firefox", program_files_x86 / "Mozilla Firefox" / "firefox.exe"),
            ("Firefox", "firefox", local_app_data / "Mozilla Firefox" / "firefox.exe"),
        ]

    seen: set[Path] = set()
    found: list[tuple[str, str, Path]] = []
    for name, engine, path in raw_candidates:
        normalized = path.resolve() if path.exists() else path
        if normalized in seen:
            continue
        seen.add(normalized)
        if path.exists():
            found.append((name, engine, path))
    return found


def find_browser(settings: Settings) -> tuple[str, str | None, str]:
    configured = settings.browser_path.strip()
    if configured:
        configured_path = Path(configured)
        configured_executable = app_bundle_executable(configured_path) or configured_path
        if configured_executable.exists():
            return (
                browser_engine_for_path(configured_executable),
                str(configured_executable),
                f"手动指定浏览器：{configured_executable}",
            )
        print(f"配置的浏览器路径不存在，改为自动查找：{configured_path}")

    candidates = browser_candidates()
    if candidates:
        name, engine, path = candidates[0]
        return engine, str(path), f"自动使用浏览器：{name} ({path})"

    return "chromium", None, "未找到已安装浏览器，尝试使用 Playwright 自带浏览器。"


def safe_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def count_to_w(count: int | None) -> float | None:
    if count is None:
        return None
    return round(count / 10000, 2)


def publish_time_from_timestamp(timestamp: Any) -> str:
    ts = safe_int(timestamp)
    if ts is None or ts <= 0:
        return ""
    return datetime.fromtimestamp(ts, CHINA_TZ).strftime("%Y-%m-%d %H:%M:%S")


def compact_number_to_int(text: str) -> int | None:
    if not text:
        return None
    cleaned = text.replace(",", "").strip()
    if "抢首评" in cleaned or cleaned in {"分享", "收藏", "点赞", "-"}:
        return 0
    match = re.search(r"(\d+(?:\.\d+)?)\s*([万wW千kK]?)", cleaned)
    if not match:
        return None
    value = float(match.group(1))
    unit = match.group(2)
    if unit in {"万", "w", "W"}:
        value *= 10000
    elif unit in {"千", "k", "K"}:
        value *= 1000
    return int(value)


def parse_fans_from_text(body_text: str) -> int | None:
    match = re.search(r"粉丝\s*([0-9.,]+(?:万|w|W)?)", body_text)
    if not match:
        return None
    return compact_number_to_int(match.group(1))


def get_aweme_detail(payload: dict[str, Any] | None) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    detail = payload.get("aweme_detail") or payload.get("aweme") or {}
    return detail if isinstance(detail, dict) else {}


def get_filter_detail_remark(payload: dict[str, Any] | None) -> str:
    if not isinstance(payload, dict):
        return ""
    filter_detail = payload.get("filter_detail")
    if not isinstance(filter_detail, dict):
        return ""
    message = filter_detail.get("detail_msg") or filter_detail.get("notice") or "作品不可观看"
    return str(message)


def register_detail_listener(page) -> asyncio.Future[dict[str, Any] | None]:
    loop = asyncio.get_running_loop()
    detail_future: asyncio.Future[dict[str, Any] | None] = loop.create_future()

    async def handle_response(response):
        if "/aweme/v1/web/aweme/detail/" not in response.url or detail_future.done():
            return
        try:
            detail_future.set_result(await response.json())
        except Exception:
            if not detail_future.done():
                detail_future.set_result(None)

    page.on("response", lambda response: asyncio.create_task(handle_response(response)))
    return detail_future


async def wait_for_detail(detail_future: asyncio.Future[dict[str, Any] | None]) -> dict[str, Any] | None:
    try:
        return await asyncio.wait_for(detail_future, timeout=20)
    except asyncio.TimeoutError:
        return None


async def goto_video_page(page, url: str) -> dict[str, Any] | None:
    detail_future = register_detail_listener(page)
    await page.goto(url, wait_until="domcontentloaded", timeout=45000)
    try:
        await page.wait_for_selector("video", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    detail = await wait_for_detail(detail_future)
    await page.wait_for_timeout(1200)
    return detail


async def get_body_text(page) -> str:
    try:
        return await page.locator("body").inner_text(timeout=5000)
    except Exception:
        return ""


async def dismiss_popups(page) -> None:
    for _ in range(3):
        try:
            await page.keyboard.press("Escape")
        except Exception:
            pass

        try:
            if await page.get_by_text("登录后免费畅享高清视频").count() > 0:
                await page.mouse.click(962, 158)
                await page.wait_for_timeout(500)
        except Exception:
            pass

        try:
            if await page.get_by_text("打开声音").count() > 0:
                await page.mouse.click(738, 459)
                await page.wait_for_timeout(250)
        except Exception:
            pass


async def seek_video(page, seconds: float) -> bool:
    try:
        return bool(
            await page.evaluate(
                """
                (target) => new Promise((resolve) => {
                    const video = document.querySelector('video');
                    if (!video) {
                        resolve(false);
                        return;
                    }
                    let finished = false;
                    const done = () => {
                        if (finished) return;
                        finished = true;
                        video.removeEventListener('seeked', done);
                        resolve(true);
                    };
                    video.muted = true;
                    video.pause();
                    video.addEventListener('seeked', done, { once: true });
                    const duration = Number.isFinite(video.duration) && video.duration > 0 ? video.duration : target;
                    video.currentTime = Math.max(0, Math.min(target, Math.max(0, duration - 0.25)));
                    setTimeout(done, 1400);
                })
                """,
                seconds,
            )
        )
    except Exception:
        return False


async def video_box(page) -> dict[str, float] | None:
    try:
        box = await page.evaluate(
            """
            () => {
                const videos = Array.from(document.querySelectorAll('video'));
                const boxes = videos.map((video) => {
                    const rect = video.getBoundingClientRect();
                    const width = Math.max(0, Math.min(rect.right, window.innerWidth) - Math.max(rect.left, 0));
                    const height = Math.max(0, Math.min(rect.bottom, window.innerHeight) - Math.max(rect.top, 0));
                    return { x: rect.left, y: rect.top, width: rect.width, height: rect.height, visibleArea: width * height };
                }).filter((box) => box.visibleArea > 1000);
                boxes.sort((a, b) => b.visibleArea - a.visibleArea);
                return boxes[0] || null;
            }
            """
        )
    except Exception:
        return None
    if not box or box.get("width", 0) <= 0 or box.get("height", 0) <= 0:
        return None
    return box


def scan_times(duration_ms: int | None) -> list[float]:
    if duration_ms and duration_ms > 3000:
        duration = duration_ms / 1000
        raw = [
            duration * 0.08,
            duration * 0.18,
            duration * 0.30,
            duration * 0.44,
            duration * 0.58,
            duration * 0.72,
            duration * 0.86,
        ]
        return sorted({round(max(0.5, min(duration - 0.5, t)), 2) for t in raw})
    return [1.0, 3.0, 6.0, 10.0, 15.0, 22.0, 30.0]


def png_to_bgr(png_bytes: bytes) -> np.ndarray:
    return cv2.imdecode(np.frombuffer(png_bytes, np.uint8), cv2.IMREAD_COLOR)


async def video_frame_has_face(
    page,
    box: dict[str, float],
    detector: FaceDetector,
    settings: Settings,
) -> bool:
    clip = {
        "x": max(0, box["x"]),
        "y": max(0, box["y"]),
        "width": max(1, min(box["width"], settings.viewport_width - max(0, box["x"]))),
        "height": max(1, min(box["height"], settings.viewport_height - max(0, box["y"]))),
    }
    png = await page.screenshot(clip=clip, full_page=False)
    return detector.has_face(png_to_bgr(png))


def adjust_crop_bounds(
    left: int,
    top: int,
    right: int,
    bottom: int,
    image_width: int,
    image_height: int,
    settings: Settings,
) -> tuple[int, int, int, int]:
    left = clamp_int(left, 0, image_width - 1)
    top = clamp_int(top, 0, image_height - 1)
    right = clamp_int(right, left + 1, image_width)
    bottom = clamp_int(bottom, top + 1, image_height)

    current_width = right - left
    if settings.crop_min_width > current_width:
        missing = settings.crop_min_width - current_width
        right = min(image_width, right + missing)
        if right - left < settings.crop_min_width:
            left = max(0, right - settings.crop_min_width)

    if settings.crop_max_width and right - left > settings.crop_max_width:
        right = min(image_width, left + settings.crop_max_width)
    if settings.crop_max_height and bottom - top > settings.crop_max_height:
        bottom = min(image_height, top + settings.crop_max_height)

    if left == 0 and top == 0 and right == image_width and bottom == image_height:
        right = max(1, image_width - 80)
        bottom = max(1, image_height - 30)
    return left, top, right, bottom


async def save_cropped_page_screenshot(
    page,
    box: dict[str, float] | None,
    stem: str,
    settings: Settings,
) -> Path:
    SCREENSHOT_DIR.mkdir(exist_ok=True)
    full_png = await page.screenshot(full_page=False)
    image = PILImage.open(io.BytesIO(full_png)).convert("RGB")

    if box:
        left = int(box["x"] - settings.crop_left_extra)
        top = int(box["y"] - settings.crop_top_extra)
        right = int(box["x"] + box["width"] + settings.crop_right_extra)
        bottom = int(box["y"] + box["height"] + settings.crop_bottom_extra)
    else:
        left, top = 0, 0
        right = min(image.width, settings.crop_min_width)
        bottom = min(image.height, settings.crop_max_height or image.height - 30)
    left, top, right, bottom = adjust_crop_bounds(
        left,
        top,
        right,
        bottom,
        image.width,
        image.height,
        settings,
    )

    cropped = image.crop((left, top, right, bottom))
    output = SCREENSHOT_DIR / f"{stem}.jpg"
    cropped.save(output, "JPEG", quality=88, optimize=True)
    return output


async def capture_screenshot_with_face(
    page,
    video_id: str,
    duration_ms: int | None,
    detector: FaceDetector,
    settings: Settings,
) -> tuple[Path | None, bool, str]:
    await dismiss_popups(page)
    box = await video_box(page)
    if not box:
        try:
            path = await save_cropped_page_screenshot(page, None, video_id, settings)
            return path, False, "未找到视频元素"
        except Exception as exc:
            return None, False, f"截图失败：{exc}"

    face_found = False
    for seconds in scan_times(duration_ms):
        await seek_video(page, seconds)
        await dismiss_popups(page)
        await page.wait_for_timeout(350)
        try:
            if await video_frame_has_face(page, box, detector, settings):
                face_found = True
                break
        except Exception:
            continue

    try:
        path = await save_cropped_page_screenshot(page, box, video_id, settings)
    except Exception as exc:
        return None, face_found, f"截图失败：{exc}"

    if face_found:
        return path, True, ""
    return path, False, "未检测到人脸"


def result_from_detail(
    item: InputItem,
    detail_payload: dict[str, Any] | None,
    body_text: str,
    old_contacts: dict[str, str],
) -> CollectResult:
    aweme = get_aweme_detail(detail_payload)
    author = aweme.get("author") or {}
    stats = aweme.get("statistics") or {}
    follower_count = safe_int(author.get("follower_count"))
    if follower_count is None:
        follower_count = parse_fans_from_text(body_text)

    return CollectResult(
        contact=old_contacts.get(item.url, ""),
        nickname=str(author.get("nickname") or ""),
        account_id=str(author.get("unique_id") or author.get("short_id") or author.get("uid") or ""),
        fans_w=count_to_w(follower_count),
        video_link=item.url,
        digg_count=safe_int(stats.get("digg_count")),
        comment_count=safe_int(stats.get("comment_count")),
        share_count=safe_int(stats.get("share_count")),
        publish_time=publish_time_from_timestamp(aweme.get("create_time")),
    )


def video_id_from_page(url: str, detail_payload: dict[str, Any] | None, fallback_url: str) -> str:
    aweme = get_aweme_detail(detail_payload)
    aweme_id = aweme.get("aweme_id")
    if aweme_id:
        return str(aweme_id)
    match = re.search(r"/video/(\d+)", url)
    if match:
        return match.group(1)
    return hashlib.md5(fallback_url.encode("utf-8")).hexdigest()[:16]


async def collect_one(
    context,
    item: InputItem,
    detector: FaceDetector,
    old_contacts: dict[str, str],
    settings: Settings,
) -> CollectResult:
    page = await context.new_page()
    page.set_default_timeout(20000)
    try:
        target_url = item.page_url or item.url
        detail = await goto_video_page(page, target_url)
        body_text = await get_body_text(page)
        filter_remark = get_filter_detail_remark(detail)
        if filter_remark:
            return CollectResult(
                contact=old_contacts.get(item.url, ""),
                video_link=item.url,
                remark=filter_remark,
            )
        result = result_from_detail(item, detail, body_text, old_contacts)
        video_id = video_id_from_page(page.url, detail, item.url)
        duration_ms = safe_int(get_aweme_detail(detail).get("duration"))
        screenshot_path, face_found, screenshot_remark = await capture_screenshot_with_face(
            page,
            video_id,
            duration_ms,
            detector,
            settings,
        )
        result.screenshot_path = screenshot_path
        remarks = []
        title = ""
        if not result.nickname:
            title = await page.title()
            result.nickname = ""
        captcha_or_failed_page = any(
            marker in f"{title}\n{body_text}"
            for marker in ("验证码", "中间页", "安全验证")
        )
        if not detail:
            if captcha_or_failed_page:
                remarks.append("验证码/加载失败")
            else:
                remarks.append("未获取到结构化数据")
        if screenshot_remark:
            remarks.append(screenshot_remark)
        if not result.nickname and title and not captcha_or_failed_page:
            remarks.append(f"仅获取到页面标题：{title[:60]}")
        result.remark = "；".join(remarks)
        return result
    except Exception as exc:
        return CollectResult(
            contact=old_contacts.get(item.url, ""),
            video_link=item.url,
            remark=f"采集失败：{exc}",
        )
    finally:
        await page.close()


async def collect_all(
    items: list[InputItem],
    detector: FaceDetector,
    old_contacts: dict[str, str],
    max_concurrent: int,
    settings: Settings,
) -> list[CollectResult]:
    browser_engine, browser_path, browser_message = find_browser(settings)
    print(browser_message)
    launch_options: dict[str, Any] = {"headless": True}
    if browser_engine == "chromium":
        launch_options["args"] = [
            "--disable-blink-features=AutomationControlled",
            "--autoplay-policy=no-user-gesture-required",
            "--disable-dev-shm-usage",
        ]
    if browser_path:
        launch_options["executable_path"] = browser_path

    async with async_playwright() as playwright:
        browser_launcher = playwright.firefox if browser_engine == "firefox" else playwright.chromium
        browser = await browser_launcher.launch(**launch_options)
        semaphore = asyncio.Semaphore(max_concurrent)

        async def run(item: InputItem) -> CollectResult:
            async with semaphore:
                print(f"采集第 {item.row} 行：{item.url}")
                context = await browser.new_context(
                    viewport={"width": settings.viewport_width, "height": settings.viewport_height},
                    user_agent=USER_AGENT,
                    locale="zh-CN",
                    timezone_id="Asia/Shanghai",
                    device_scale_factor=1,
                )
                await context.add_init_script(
                    "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
                )
                try:
                    return await collect_one(context, item, detector, old_contacts, settings)
                finally:
                    await context.close()

        try:
            return await asyncio.gather(*(run(item) for item in items))
        finally:
            await browser.close()


def should_retry(result: CollectResult) -> bool:
    terminal_markers = ("作品权限或已被删除", "无法观看", "作品不见了", "作品不可观看")
    if any(marker in (result.remark or "") for marker in terminal_markers):
        return False
    retry_markers = ("验证码", "未获取到结构化数据", "未找到视频元素", "采集失败")
    if any(marker in (result.remark or "") for marker in retry_markers):
        return True
    return not (result.nickname and result.publish_time)


def result_quality(result: CollectResult) -> int:
    score = 0
    if result.nickname:
        score += 3
    if result.account_id:
        score += 2
    if result.publish_time:
        score += 2
    if result.screenshot_path and result.screenshot_path.exists():
        score += 2
    if not should_retry(result):
        score += 3
    return score


def should_replace_with_retry(current: CollectResult, retry_result: CollectResult) -> bool:
    return result_quality(retry_result) >= result_quality(current)


def build_output_workbook(results: list[CollectResult], settings: Settings) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "采集结果"
    ws.append(OUTPUT_HEADERS)

    header_fill = PatternFill("solid", fgColor="44546A")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = [16, 24, 20, 12, 48, 10, 10, 10, 22, 46, 34]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    for result in results:
        row_index = ws.max_row + 1
        ws.append(
            [
                result.contact,
                result.nickname,
                result.account_id,
                result.fans_w,
                result.video_link,
                result.digg_count,
                result.comment_count,
                result.share_count,
                result.publish_time,
                "",
                result.remark,
            ]
        )
        ws.row_dimensions[row_index].height = 165
        for col in range(1, len(OUTPUT_HEADERS) + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.cell(row=row_index, column=4).number_format = "0.00"
        for col in (6, 7, 8):
            ws.cell(row=row_index, column=col).number_format = "0"

        if result.screenshot_path and result.screenshot_path.exists():
            try:
                image = ExcelImage(str(result.screenshot_path))
                max_width = settings.excel_image_max_width
                max_height = settings.excel_image_max_height
                scale = min(max_width / image.width, max_height / image.height, 1)
                image.width = int(image.width * scale)
                image.height = int(image.height * scale)
                ws.add_image(image, f"J{row_index}")
            except Exception:
                current = ws.cell(row=row_index, column=11).value or ""
                ws.cell(row=row_index, column=11).value = (current + "；截图嵌入失败").strip("；")

    OUTPUT_XLSX.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT_XLSX)


def main() -> int:
    settings = load_settings()
    created = ensure_input_workbook()
    items = read_input_items()
    if not items:
        if created:
            print(f"已创建输入模板：{INPUT_XLSX}")
        print("未发现抖音链接。请把抖音分享链接填到“待采集视频链接.xlsx”的 A 列后重新运行。")
        return 0

    old_contacts = read_existing_contacts()
    print(f"发现 {len(items)} 条抖音链接，开始采集。")
    print(f"配置文件：{CONFIG_TXT}")
    print(
        "当前配置："
        f"一次性采集数量={settings.max_concurrent}，"
        f"截图左/右/上/下扩展={settings.crop_left_extra}/"
            f"{settings.crop_right_extra}/{settings.crop_top_extra}/{settings.crop_bottom_extra}"
    )
    items = prepare_input_items(items)
    model = ensure_yunet_model()
    if model:
        print("人脸检测：使用 YuNet 模型。")
    else:
        print("人脸检测：YuNet 模型下载失败，使用 OpenCV 内置检测器兜底。")
    detector = FaceDetector(model)

    results = asyncio.run(
        collect_all(items, detector, old_contacts, settings.max_concurrent, settings)
    )
    for retry_round in range(1, MAX_RETRY_ROUNDS + 1):
        retry_indexes = [idx for idx, result in enumerate(results) if should_retry(result)]
        if not retry_indexes:
            break
        print(f"有 {len(retry_indexes)} 条疑似被验证码或加载失败，第 {retry_round} 次降速重试。")
        retry_items = [items[idx] for idx in retry_indexes]
        retry_results = asyncio.run(
            collect_all(retry_items, detector, old_contacts, settings.retry_concurrent, settings)
        )
        for idx, retry_result in zip(retry_indexes, retry_results):
            if should_replace_with_retry(results[idx], retry_result):
                results[idx] = retry_result
    build_output_workbook(results, settings)
    final_failures = sum(1 for result in results if should_retry(result))
    final_unavailable = sum(
        1
        for result in results
        if any(marker in (result.remark or "") for marker in ("作品权限或已被删除", "无法观看", "作品不见了", "作品不可观看"))
    )
    final_successes = len(results) - final_failures - final_unavailable
    print(
        f"最终结果：成功 {final_successes} 条，"
        f"作品不可采集 {final_unavailable} 条，失败/需人工复查 {final_failures} 条。"
    )
    print(f"采集完成：{OUTPUT_XLSX}")
    print(f"截图目录：{SCREENSHOT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
